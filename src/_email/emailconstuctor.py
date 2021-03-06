import datetime
from openpyxl import load_workbook
from dataCollector.datacollector import DataCollector
from PLOTS.plots import main as plot_main
from PLOTS.uploadimg import plots_urls


def get_email_data():

    #get data
    data = DataCollector()
    print('Collecting DRS data...')
    drs_total = data.getDrsNumber()
    mias_total = data.getIsMia()
    nome_modelo = data.getModels(IsSorted=True)

    #get most requested and top 5
    fabrics =  data.getFabrics(IsSorted=True)
    most_req_fabric = fabrics[0][0]

    top_5 = []

    for x, y in enumerate(fabrics):
        if x < 6 and x > 0:
            top_5.append(y[0])

    top_5_edit = f'{top_5[0]} | {top_5[1]} | {top_5[2]} | {top_5[3]} | {top_5[4]}'

    #get name of most requested fabric
    file = load_workbook(filename=r'./assets/TC_02.xlsx')
    old_file = load_workbook(filename=r'./assets/tc_oldlist.xlsx')
    sheet = file['Sheet1']
    sheet_old = old_file['Sheet1']

    fabric_name = ''
    x = 1
    while True:
        controller = sheet[f'A{x}'].value
        if controller == None:
            break
        cod = sheet[f'C{x}'].value
        if cod == most_req_fabric.upper():
            name = sheet[f'D{x}'].value

            if len(name) <= 0:
                j = 4
                while True:
                    cont = sheet_old[f'A{j}'].value
                    #print(cont)
                    if cont == None:
                        break
                    value = sheet_old[f'C{j}'].value
                    if value == cod:
                        name_tretmant = sheet_old[f'B{j}'].value
                        fabric_name = name_tretmant.replace('Tecido', '').replace('s/tratamento', '').strip().capitalize()
                        break
                    j += 1
                break
            else:
                fabric_name = name
                break
        x += 1
    
    file.close()

    #create plots
    plot_main()
    plots = plots_urls()

    email_data = {
        'most_req_fabric':most_req_fabric,
        'top_5':top_5_edit,
        'fabric_name':fabric_name,
        'drs_total':drs_total,
        'mias_total':mias_total,
        'nome_modelo':nome_modelo,
        'plot_url':plots
    }
    print('Data collection - DONE')
    return email_data


def gen_email():

    #dates
    now = datetime.datetime.now()
    year = now.year
    month = {
        1:'Jan',
        2:'Fev',
        3:'Mar',
        4:'Apr',
        5:'May',
        6:'Jun',
        7:'Jul',
        8:'Aug',
        9:'Sep',
        10:'Oct',
        11:'Nov',
        12:'Dec'
    }

    date = f'{month[now.month - 1]} {year}'

    data = get_email_data()


    #build html

    html =f'''

<!DOCTYPE html PUBLIC "-//W3C//DTD XHTML 1.0 Transitional//EN" "http://www.w3.org/TR/xhtml1/DTD/xhtml1-transitional.dtd">
<html xmlns="http://www.w3.org/1999/xhtml" xmlns:o="urn:schemas-microsoft-com:office:office">

<head>
    <meta content="width=device-width, initial-scale=1" name="viewport">
    <meta charset="UTF-8">
    <meta name="x-apple-disable-message-reformatting">
    <meta http-equiv="X-UA-Compatible" content="IE=edge">
    <meta content="telephone=no" name="format-detection">
    <title></title>
    <!--[if (mso 16)]>
    <style type="text/css">
    a {{text-decoration: none;}}
    </style>
    <![endif]-->
    <!--[if gte mso 9]><style>sup {{ font-size: 100% !important; }}</style><![endif]-->
    <!--[if gte mso 9]>
<xml>
    <o:OfficeDocumentSettings>
    <o:AllowPNG></o:AllowPNG>
    <o:PixelsPerInch>96</o:PixelsPerInch>
    </o:OfficeDocumentSettings>
</xml>
<![endif]-->
    <!--[if !mso]><!-- -->
    <link href="https://fonts.googleapis.com/css?family=Roboto:400,400i,700,700i" rel="stylesheet">
    <!--<![endif]-->
    <style>


        /* CONFIG STYLES Please do not delete and edit CSS styles below */
/* IMPORTANT THIS STYLES MUST BE ON FINAL EMAIL */
#outlook a {{
    padding: 0;
}}

.es-button {{
    mso-style-priority: 100 !important;
    text-decoration: none !important;
}}

a[x-apple-data-detectors] {{
    color: inherit !important;
    text-decoration: none !important;
    font-size: inherit !important;
    font-family: inherit !important;
    font-weight: inherit !important;
    line-height: inherit !important;
}}

.es-desk-hidden {{
    display: none;
    float: left;
    overflow: hidden;
    width: 0;
    max-height: 0;
    line-height: 0;
    mso-hide: all;
}}

[data-ogsb] .es-button {{
    border-width: 0 !important;
    padding: 10px 20px 10px 20px !important;
}}

/*
END OF IMPORTANT
*/
body {{
    width: 100%;
    font-family: arial, 'helvetica neue', helvetica, sans-serif;
    -webkit-text-size-adjust: 100%;
    -ms-text-size-adjust: 100%;
}}

table {{
    mso-table-lspace: 0pt;
    mso-table-rspace: 0pt;
    border-collapse: collapse;
    border-spacing: 0px;
}}

table td,
body,
.es-wrapper {{
    padding: 0;
    Margin: 0;
}}

.es-content,
.es-header,
.es-footer {{
    table-layout: fixed !important;
    width: 100%;
}}

img {{
    display: block;
    border: 0;
    outline: none;
    text-decoration: none;
    -ms-interpolation-mode: bicubic;
}}

p,
hr {{
    Margin: 0;
}}

h1,
h2,
h3,
h4,
h5 {{
    Margin: 0;
    line-height: 120%;
    mso-line-height-rule: exactly;
    font-family: roboto, 'helvetica neue', helvetica, arial, sans-serif;
}}

p,
ul li,
ol li,
a {{
    -webkit-text-size-adjust: none;
    -ms-text-size-adjust: none;
    mso-line-height-rule: exactly;
}}

.es-left {{
    float: left;
}}

.es-right {{
    float: right;
}}

.es-p5 {{
    padding: 5px;
}}

.es-p5t {{
    padding-top: 5px;
}}

.es-p5b {{
    padding-bottom: 5px;
}}

.es-p5l {{
    padding-left: 5px;
}}

.es-p5r {{
    padding-right: 5px;
}}

.es-p10 {{
    padding: 10px;
}}

.es-p10t {{
    padding-top: 10px;
}}

.es-p10b {{
    padding-bottom: 10px;
}}

.es-p10l {{
    padding-left: 10px;
}}

.es-p10r {{
    padding-right: 10px;
}}

.es-p15 {{
    padding: 15px;
}}

.es-p15t {{
    padding-top: 15px;
}}

.es-p15b {{
    padding-bottom: 15px;
}}

.es-p15l {{
    padding-left: 15px;
}}

.es-p15r {{
    padding-right: 15px;
}}

.es-p20 {{
    padding: 20px;
}}

.es-p20t {{
    padding-top: 20px;
}}

.es-p20b {{
    padding-bottom: 20px;
}}

.es-p20l {{
    padding-left: 20px;
}}

.es-p20r {{
    padding-right: 20px;
}}

.es-p25 {{
    padding: 25px;
}}

.es-p25t {{
    padding-top: 25px;
}}

.es-p25b {{
    padding-bottom: 25px;
}}

.es-p25l {{
    padding-left: 25px;
}}

.es-p25r {{
    padding-right: 25px;
}}

.es-p30 {{
    padding: 30px;
}}

.es-p30t {{
    padding-top: 30px;
}}

.es-p30b {{
    padding-bottom: 30px;
}}

.es-p30l {{
    padding-left: 30px;
}}

.es-p30r {{
    padding-right: 30px;
}}

.es-p35 {{
    padding: 35px;
}}

.es-p35t {{
    padding-top: 35px;
}}

.es-p35b {{
    padding-bottom: 35px;
}}

.es-p35l {{
    padding-left: 35px;
}}

.es-p35r {{
    padding-right: 35px;
}}

.es-p40 {{
    padding: 40px;
}}

.es-p40t {{
    padding-top: 40px;
}}

.es-p40b {{
    padding-bottom: 40px;
}}

.es-p40l {{
    padding-left: 40px;
}}

.es-p40r {{
    padding-right: 40px;
}}

.es-menu td {{
    border: 0;
}}

.es-menu td a img {{
    display: inline-block !important;
    vertical-align: middle;
}}

/*
END CONFIG STYLES
*/
s {{
    text-decoration: line-through;
}}

p,
ul li,
ol li {{
    font-family: arial, 'helvetica neue', helvetica, sans-serif;
    line-height: 150%;
}}

ul li,
ol li {{
    Margin-bottom: 15px;
}}

a {{
    text-decoration: underline;
}}

.es-menu td a {{
    text-decoration: none;
    display: block;
}}

.es-wrapper {{
    width: 100%;
    height: 100%;
    
    background-repeat: repeat;
    background-position: center top;
}}

.es-wrapper-color {{
    background-color: #f6f6f6;
}}

.es-header {{
    background-color: transparent;
    
    background-repeat: repeat;
    background-position: center top;
}}

.es-header-body {{
    background-color: #ffffff;
}}

.es-header-body p,
.es-header-body ul li,
.es-header-body ol li {{
    color: #333333;
    font-size: 14px;
}}

.es-header-body a {{
    color: #1376c8;
    font-size: 14px;
}}

.es-content-body {{
    background-color: #ffffff;
}}

.es-content-body p,
.es-content-body ul li,
.es-content-body ol li {{
    color: #333333;
    font-size: 14px;
}}

.es-content-body a {{
    color: #2cb543;
    font-size: 14px;
}}

.es-footer {{
    background-color: transparent;
    
    background-repeat: repeat;
    background-position: center top;
}}

.es-footer-body {{
    background-color: #ffffff;
}}

.es-footer-body p,
.es-footer-body ul li,
.es-footer-body ol li {{
    color: #333333;
    font-size: 14px;
}}

.es-footer-body a {{
    color: #ffffff;
    font-size: 14px;
}}

.es-infoblock,
.es-infoblock p,
.es-infoblock ul li,
.es-infoblock ol li {{
    line-height: 120%;
    font-size: 12px;
    color: #cccccc;
}}

.es-infoblock a {{
    font-size: 12px;
    color: #cccccc;
}}

h1 {{
    font-size: 30px;
    font-style: normal;
    font-weight: normal;
    color: #333333;
}}

h2 {{
    font-size: 24px;
    font-style: normal;
    font-weight: normal;
    color: #333333;
}}

h3 {{
    font-size: 20px;
    font-style: normal;
    font-weight: normal;
    color: #333333;
}}

.es-header-body h1 a,
.es-content-body h1 a,
.es-footer-body h1 a {{
    font-size: 30px;
}}

.es-header-body h2 a,
.es-content-body h2 a,
.es-footer-body h2 a {{
    font-size: 24px;
}}

.es-header-body h3 a,
.es-content-body h3 a,
.es-footer-body h3 a {{
    font-size: 20px;
}}

a.es-button,
button.es-button {{
    border-style: solid;
    border-color: #31cb4b;
    border-width: 10px 20px 10px 20px;
    display: inline-block;
    background: #31cb4b;
    border-radius: 30px;
    font-size: 18px;
    font-family: arial, 'helvetica neue', helvetica, sans-serif;
    font-weight: normal;
    font-style: normal;
    line-height: 120%;
    color: #ffffff;
    text-decoration: none;
    width: auto;
    text-align: center;
}}

.es-button-border {{
    border-style: solid solid solid solid;
    border-color: #2cb543 #2cb543 #2cb543 #2cb543;
    background: #2cb543;
    border-width: 0px 0px 2px 0px;
    display: inline-block;
    border-radius: 30px;
    width: auto;
}}

/* RESPONSIVE STYLES Please do not delete and edit CSS styles below. If you don't need responsive layout, please delete this section. */
@media only screen and (max-width: 600px) {{

    p,
    ul li,
    ol li,
    a {{
        line-height: 150% !important;
    }}

    h1 {{
        font-size: 30px !important;
        text-align: center;
        line-height: 120% !important;
    }}

    h2 {{
        font-size: 26px !important;
        text-align: center;
        line-height: 120% !important;
    }}

    h3 {{
        font-size: 20px !important;
        text-align: center;
        line-height: 120% !important;
    }}

    .es-header-body h1 a,
    .es-content-body h1 a,
    .es-footer-body h1 a {{
        font-size: 30px !important;
    }}

    .es-header-body h2 a,
    .es-content-body h2 a,
    .es-footer-body h2 a {{
        font-size: 26px !important;
    }}

    .es-header-body h3 a,
    .es-content-body h3 a,
    .es-footer-body h3 a {{
        font-size: 20px !important;
    }}

    .es-menu td a {{
        font-size: 16px !important;
    }}

    .es-header-body p,
    .es-header-body ul li,
    .es-header-body ol li,
    .es-header-body a {{
        font-size: 16px !important;
    }}

    .es-content-body p,
    .es-content-body ul li,
    .es-content-body ol li,
    .es-content-body a {{
        font-size: 16px !important;
    }}

    .es-footer-body p,
    .es-footer-body ul li,
    .es-footer-body ol li,
    .es-footer-body a {{
        font-size: 16px !important;
    }}

    .es-infoblock p,
    .es-infoblock ul li,
    .es-infoblock ol li,
    .es-infoblock a {{
        font-size: 12px !important;
    }}

    *[class="gmail-fix"] {{
        display: none !important;
    }}

    .es-m-txt-c,
    .es-m-txt-c h1,
    .es-m-txt-c h2,
    .es-m-txt-c h3 {{
        text-align: center !important;
    }}

    .es-m-txt-r,
    .es-m-txt-r h1,
    .es-m-txt-r h2,
    .es-m-txt-r h3 {{
        text-align: right !important;
    }}

    .es-m-txt-l,
    .es-m-txt-l h1,
    .es-m-txt-l h2,
    .es-m-txt-l h3 {{
        text-align: left !important;
    }}

    .es-m-txt-r img,
    .es-m-txt-c img,
    .es-m-txt-l img {{
        display: inline !important;
    }}

    .es-button-border {{
        display: block !important;
    }}

    a.es-button,
    button.es-button {{
        font-size: 20px !important;
        display: block !important;
        border-left-width: 0px !important;
        border-right-width: 0px !important;
    }}

    .es-adaptive table,
    .es-left,
    .es-right {{
        width: 100% !important;
    }}

    .es-content table,
    .es-header table,
    .es-footer table,
    .es-content,
    .es-footer,
    .es-header {{
        width: 100% !important;
        max-width: 600px !important;
    }}

    .es-adapt-td {{
        display: block !important;
        width: 100% !important;
    }}

    .adapt-img {{
        width: 100% !important;
        height: auto !important;
    }}

    .es-m-p0 {{
        padding: 0 !important;
    }}

    .es-m-p0r {{
        padding-right: 0 !important;
    }}

    .es-m-p0l {{
        padding-left: 0 !important;
    }}

    .es-m-p0t {{
        padding-top: 0 !important;
    }}

    .es-m-p0b {{
        padding-bottom: 0 !important;
    }}

    .es-m-p20b {{
        padding-bottom: 20px !important;
    }}

    .es-mobile-hidden,
    .es-hidden {{
        display: none !important;
    }}

    tr.es-desk-hidden,
    td.es-desk-hidden,
    table.es-desk-hidden {{
        width: auto !important;
        overflow: visible !important;
        float: none !important;
        max-height: inherit !important;
        line-height: inherit !important;
    }}

    tr.es-desk-hidden {{
        display: table-row !important;
    }}

    table.es-desk-hidden {{
        display: table !important;
    }}

    td.es-desk-menu-hidden {{
        display: table-cell !important;
    }}

    .es-menu td {{
        width: 1% !important;
    }}

    table.es-table-not-adapt,
    .esd-block-html table {{
        width: auto !important;
    }}

    table.es-social {{
        display: inline-block !important;
    }}

    table.es-social td {{
        display: inline-block !important;
    }}

    .es-m-p5 {{
        padding: 5px !important;
    }}

    .es-m-p5t {{
        padding-top: 5px !important;
    }}

    .es-m-p5b {{
        padding-bottom: 5px !important;
    }}

    .es-m-p5r {{
        padding-right: 5px !important;
    }}

    .es-m-p5l {{
        padding-left: 5px !important;
    }}

    .es-m-p10 {{
        padding: 10px !important;
    }}

    .es-m-p10t {{
        padding-top: 10px !important;
    }}

    .es-m-p10b {{
        padding-bottom: 10px !important;
    }}

    .es-m-p10r {{
        padding-right: 10px !important;
    }}

    .es-m-p10l {{
        padding-left: 10px !important;
    }}

    .es-m-p15 {{
        padding: 15px !important;
    }}

    .es-m-p15t {{
        padding-top: 15px !important;
    }}

    .es-m-p15b {{
        padding-bottom: 15px !important;
    }}

    .es-m-p15r {{
        padding-right: 15px !important;
    }}

    .es-m-p15l {{
        padding-left: 15px !important;
    }}

    .es-m-p20 {{
        padding: 20px !important;
    }}

    .es-m-p20t {{
        padding-top: 20px !important;
    }}

    .es-m-p20r {{
        padding-right: 20px !important;
    }}

    .es-m-p20l {{
        padding-left: 20px !important;
    }}

    .es-m-p25 {{
        padding: 25px !important;
    }}

    .es-m-p25t {{
        padding-top: 25px !important;
    }}

    .es-m-p25b {{
        padding-bottom: 25px !important;
    }}

    .es-m-p25r {{
        padding-right: 25px !important;
    }}

    .es-m-p25l {{
        padding-left: 25px !important;
    }}

    .es-m-p30 {{
        padding: 30px !important;
    }}

    .es-m-p30t {{
        padding-top: 30px !important;
    }}

    .es-m-p30b {{
        padding-bottom: 30px !important;
    }}

    .es-m-p30r {{
        padding-right: 30px !important;
    }}

    .es-m-p30l {{
        padding-left: 30px !important;
    }}

    .es-m-p35 {{
        padding: 35px !important;
    }}

    .es-m-p35t {{
        padding-top: 35px !important;
    }}

    .es-m-p35b {{
        padding-bottom: 35px !important;
    }}

    .es-m-p35r {{
        padding-right: 35px !important;
    }}

    .es-m-p35l {{
        padding-left: 35px !important;
    }}

    .es-m-p40 {{
        padding: 40px !important;
    }}

    .es-m-p40t {{
        padding-top: 40px !important;
    }}

    .es-m-p40b {{
        padding-bottom: 40px !important;
    }}

    .es-m-p40r {{
        padding-right: 40px !important;
    }}

    .es-m-p40l {{
        padding-left: 40px !important;
    }}
}}

/* END RESPONSIVE STYLES */


    </style>
</head>

<body>
    <div class="es-wrapper-color">
        <!--[if gte mso 9]>
			<v:background xmlns:v="urn:schemas-microsoft-com:vml" fill="t">
				<v:fill type="tile" color="#f6f6f6"></v:fill>
			</v:background>
		<![endif]-->
        <table class="es-wrapper" width="100%" cellspacing="0" cellpadding="0">
            <tbody>
                <tr>
                    <td class="esd-email-paddings" valign="top">
                        <table class="esd-header-popover es-header" cellspacing="0" cellpadding="0" align="center">
                            <tbody>
                                <tr>
                                    <td class="esd-stripe" align="center">
                                        <table class="es-header-body" width="650" cellspacing="0" cellpadding="0" bgcolor="#ffffff" align="center">
                                            <tbody>
                                                <tr>
                                                    <td class="esd-structure es-p20r es-p20l" align="left" bgcolor="#666666" style="background-color: #666666;">
                                                        <table cellspacing="0" cellpadding="0" width="100%">
                                                            <tbody>
                                                                <tr>
                                                                    <td class="esd-container-frame" width="610" align="left">
                                                                        <table width="100%" cellspacing="0" cellpadding="0">
                                                                            <tbody>
                                                                                <tr>
                                                                                    <td align="center" class="esd-block-text es-m-txt-c es-p40 es-m-p10">
                                                                                        <h1 style="color: #ffffff; font-size: 72px; font-family: roboto, 'helvetica neue', helvetica, arial, sans-serif; line-height: 120%;"><strong>DRS</strong></h1>
                                                                                        <h1 style="color: #ffffff; font-size: 40px; font-family: roboto, 'helvetica neue', helvetica, arial, sans-serif; line-height: 120%;">reports</h1><span style="color: #ffffff; font-family: roboto, 'helvetica neue', helvetica, arial, sans-serif; font-size: 20px; line-height: 120%;">{date}</span>
                                                                                    </td>
                                                                                </tr>
                                                                            </tbody>
                                                                        </table>
                                                                    </td>
                                                                </tr>
                                                            </tbody>
                                                        </table>
                                                    </td>
                                                </tr>
                                            </tbody>
                                        </table>
                                    </td>
                                </tr>
                            </tbody>
                        </table>
                        <table cellpadding="0" cellspacing="0" class="es-content" align="center">
                            <tbody>
                                <tr>
                                    <td class="esd-stripe" align="center">
                                        <table bgcolor="#ffffff" class="es-content-body" align="center" cellpadding="0" cellspacing="0" width="650">
                                            <tbody>
                                                <tr>
                                                    <td class="esd-structure es-p20r es-p20l" align="left">
                                                        <table cellpadding="0" cellspacing="0" width="100%">
                                                            <tbody>
                                                                <tr>
                                                                    <td width="610" class="esd-container-frame" align="center" valign="top">
                                                                        <table cellpadding="0" cellspacing="0" width="100%">
                                                                            <tbody>
                                                                                <tr>
                                                                                    <td align="center" class="esd-block-spacer es-p5" style="font-size:0">
                                                                                        <table border="0" width="100%" height="100%" cellpadding="0" cellspacing="0">
                                                                                            <tbody>
                                                                                                <tr>
                                                                                                    <td style="border-bottom: 1px solid #cccccc; background:none; height:1px; width:100%; margin:0px 0px 0px 0px;"></td>
                                                                                                </tr>
                                                                                            </tbody>
                                                                                        </table>
                                                                                    </td>
                                                                                </tr>
                                                                            </tbody>
                                                                        </table>
                                                                    </td>
                                                                </tr>
                                                            </tbody>
                                                        </table>
                                                    </td>
                                                </tr>
                                            </tbody>
                                        </table>
                                    </td>
                                </tr>
                            </tbody>
                        </table>
                        <table class="es-content" cellspacing="0" cellpadding="0" align="center">
                            <tbody>
                                <tr>
                                    <td class="esd-stripe" align="center">
                                        <table class="es-content-body" width="650" cellspacing="0" cellpadding="0" bgcolor="#ffffff" align="center">
                                            <tbody>
                                                <tr>
                                                    <td class="esd-structure es-p10t es-p20b es-p20r es-p20l" align="left">
                                                        <!--[if mso]><table width="610" cellpadding="0" cellspacing="0"><tr><td width="295" valign="top"><![endif]-->
                                                        <table cellspacing="0" cellpadding="0" align="left" class="es-left">
                                                            <tbody>
                                                                <tr>
                                                                    <td class="esd-container-frame es-m-p20b" width="295" valign="top" align="center">
                                                                        <table width="100%" cellspacing="0" cellpadding="0">
                                                                            <tbody>
                                                                                <tr>
                                                                                    <td align="center" class="esd-block-text">
                                                                                        <h1 style="font-size: 70px; line-height: 120%; color: #000000; font-family: roboto, 'helvetica neue', helvetica, arial, sans-serif;"><strong>{data['drs_total']}</strong></h1>
                                                                                        <h3 style="font-size: 70px; line-height: 120%; color: #000000; font-family: roboto, 'helvetica neue', helvetica, arial, sans-serif;"><strong><span style="font-size: 20px; line-height: 120%;">TOTAL ISSUED</span></strong></h3>
                                                                                    </td>
                                                                                </tr>
                                                                            </tbody>
                                                                        </table>
                                                                    </td>
                                                                </tr>
                                                            </tbody>
                                                        </table>
                                                        <!--[if mso]></td><td width="20"></td><td width="295" valign="top"><![endif]-->
                                                        <table cellpadding="0" cellspacing="0" class="es-right" align="right">
                                                            <tbody>
                                                                <tr>
                                                                    <td class="esd-container-frame" width="295" valign="top" align="center">
                                                                        <table width="100%" cellspacing="0" cellpadding="0">
                                                                            <tbody>
                                                                                <tr>
                                                                                    <td align="center" class="esd-block-text">
                                                                                        <h1 style="font-size: 70px; line-height: 120%; color: #000000; font-family: roboto, 'helvetica neue', helvetica, arial, sans-serif;"><b>{data['mias_total']}</b></h1>
                                                                                        <h3 style="font-size: 20px; line-height: 120%; color: #000000; font-family: roboto, 'helvetica neue', helvetica, arial, sans-serif;"><strong>MIA'S</strong></h3>
                                                                                    </td>
                                                                                </tr>
                                                                            </tbody>
                                                                        </table>
                                                                    </td>
                                                                </tr>
                                                            </tbody>
                                                        </table>
                                                        <!--[if mso]></td></tr></table><![endif]-->
                                                    </td>
                                                </tr>
                                            </tbody>
                                        </table>
                                    </td>
                                </tr>
                            </tbody>
                        </table>
                        <table cellpadding="0" cellspacing="0" class="es-content" align="center">
                            <tbody>
                                <tr>
                                    <td class="esd-stripe" align="center">
                                        <table bgcolor="#ffffff" class="es-content-body" align="center" cellpadding="0" cellspacing="0" width="650">
                                            <tbody>
                                                <tr>
                                                    <td class="esd-structure es-p20r es-p20l" align="left">
                                                        <table cellpadding="0" cellspacing="0" width="100%">
                                                            <tbody>
                                                                <tr>
                                                                    <td width="610" class="esd-container-frame" align="center" valign="top">
                                                                        <table cellpadding="0" cellspacing="0" width="100%">
                                                                            <tbody>
                                                                                <tr>
                                                                                    <td align="center" class="esd-block-spacer es-p5" style="font-size:0">
                                                                                        <table border="0" width="100%" height="100%" cellpadding="0" cellspacing="0">
                                                                                            <tbody>
                                                                                                <tr>
                                                                                                    <td style="border-bottom: 1px solid #cccccc; background:none; height:1px; width:100%; margin:0px 0px 0px 0px;"></td>
                                                                                                </tr>
                                                                                            </tbody>
                                                                                        </table>
                                                                                    </td>
                                                                                </tr>
                                                                            </tbody>
                                                                        </table>
                                                                    </td>
                                                                </tr>
                                                            </tbody>
                                                        </table>
                                                    </td>
                                                </tr>
                                            </tbody>
                                        </table>
                                    </td>
                                </tr>
                            </tbody>
                        </table>
                        <table cellpadding="0" cellspacing="0" class="es-content" align="center">
                            <tbody>
                                <tr>
                                    <td class="esd-stripe" align="center">
                                        <table bgcolor="#ffffff" class="es-content-body" align="center" cellpadding="0" cellspacing="0" width="650">
                                            <tbody>
                                                <tr>
                                                    <td class="esd-structure es-p25t es-p20b es-p20r es-p20l" align="left">
                                                        <!--[if mso]><table width="610" cellpadding="0" cellspacing="0"><tr><td width="271" valign="top"><![endif]-->
                                                        <table cellpadding="0" cellspacing="0" align="left" class="es-left">
                                                            <tbody>
                                                                <tr>
                                                                    <td width="271" class="esd-container-frame es-m-p20b" align="center" valign="top">
                                                                        <table cellpadding="0" cellspacing="0" width="100%">
                                                                            <tbody>
                                                                                <tr>
                                                                                    <td align="center" class="esd-block-text">
                                                                                        <h1 style="line-height: 120%; color: #000000;"><strong>Most requested fabric</strong></h1>
                                                                                    </td>
                                                                                </tr>
                                                                            </tbody>
                                                                        </table>
                                                                    </td>
                                                                </tr>
                                                            </tbody>
                                                        </table>
                                                        <!--[if mso]></td><td width="20"></td><td width="319" valign="top"><![endif]-->
                                                        <table cellpadding="0" cellspacing="0" class="es-right" align="right">
                                                            <tbody>
                                                                <tr>
                                                                    <td width="319" align="left" class="esd-container-frame">
                                                                        <table cellpadding="0" cellspacing="0" width="100%">
                                                                            <tbody>
                                                                                <tr>
                                                                                    <td align="center" class="esd-block-text">
                                                                                        <h1 style="font-size: 45px; color: #000000;"><strong>{data['most_req_fabric']}</strong></h1>
                                                                                    </td>
                                                                                </tr>
                                                                                <tr>
                                                                                    <td align="center" class="esd-block-text">
                                                                                        <h3 style="font-size: 15px; color: #000000;">{data['fabric_name']}</h3>
                                                                                    </td>
                                                                                </tr>
                                                                            </tbody>
                                                                        </table>
                                                                    </td>
                                                                </tr>
                                                            </tbody>
                                                        </table>
                                                        <!--[if mso]></td></tr></table><![endif]-->
                                                    </td>
                                                </tr>
                                            </tbody>
                                        </table>
                                    </td>
                                </tr>
                            </tbody>
                        </table>
                        <table cellpadding="0" cellspacing="0" class="es-content" align="center">
                            <tbody>
                                <tr>
                                    <td class="esd-stripe" align="center">
                                        <table bgcolor="#ffffff" class="es-content-body" align="center" cellpadding="0" cellspacing="0" width="650">
                                            <tbody>
                                                <tr>
                                                    <td class="esd-structure es-p20r es-p20l" align="left">
                                                        <table cellpadding="0" cellspacing="0" width="100%">
                                                            <tbody>
                                                                <tr>
                                                                    <td width="610" class="esd-container-frame" align="center" valign="top">
                                                                        <table cellpadding="0" cellspacing="0" width="100%">
                                                                            <tbody>
                                                                                <tr>
                                                                                    <td align="center" class="esd-block-spacer es-p5" style="font-size:0">
                                                                                        <table border="0" width="100%" height="100%" cellpadding="0" cellspacing="0">
                                                                                            <tbody>
                                                                                                <tr>
                                                                                                    <td style="border-bottom: 1px solid #cccccc; background:none; height:1px; width:100%; margin:0px 0px 0px 0px;"></td>
                                                                                                </tr>
                                                                                            </tbody>
                                                                                        </table>
                                                                                    </td>
                                                                                </tr>
                                                                            </tbody>
                                                                        </table>
                                                                    </td>
                                                                </tr>
                                                            </tbody>
                                                        </table>
                                                    </td>
                                                </tr>
                                            </tbody>
                                        </table>
                                    </td>
                                </tr>
                            </tbody>
                        </table>
                        <table cellpadding="0" cellspacing="0" class="es-content" align="center">
                            <tbody>
                                <tr>
                                    <td class="esd-stripe" align="center">
                                        <table bgcolor="#ffffff" class="es-content-body" align="center" cellpadding="0" cellspacing="0" width="650">
                                            <tbody>
                                                <tr>
                                                    <td class="esd-structure es-p5t es-p5b es-p20r es-p20l" align="left">
                                                        <table cellpadding="0" cellspacing="0" width="100%">
                                                            <tbody>
                                                                <tr>
                                                                    <td width="609" align="left" class="esd-container-frame">
                                                                        <table cellpadding="0" cellspacing="0" width="100%">
                                                                            <tbody>
                                                                                <tr>
                                                                                    <td align="center" class="esd-block-text es-m-txt-c">
                                                                                        <h2 style="color: #000000; line-height: 150%;"><strong>{data['top_5']}</strong></h2>
                                                                                    </td>
                                                                                </tr>
                                                                            </tbody>
                                                                        </table>
                                                                    </td>
                                                                </tr>
                                                            </tbody>
                                                        </table>
                                                    </td>
                                                </tr>
                                            </tbody>
                                        </table>
                                    </td>
                                </tr>
                            </tbody>
                        </table>
                        <table cellpadding="0" cellspacing="0" class="es-content" align="center">
                            <tbody>
                                <tr>
                                    <td class="esd-stripe" align="center">
                                        <table bgcolor="#ffffff" class="es-content-body" align="center" cellpadding="0" cellspacing="0" width="650">
                                            <tbody>
                                                <tr>
                                                    <td class="esd-structure es-p20r es-p20l" align="left">
                                                        <table cellpadding="0" cellspacing="0" width="100%">
                                                            <tbody>
                                                                <tr>
                                                                    <td width="610" class="esd-container-frame" align="center" valign="top">
                                                                        <table cellpadding="0" cellspacing="0" width="100%">
                                                                            <tbody>
                                                                                <tr>
                                                                                    <td align="center" class="esd-block-spacer es-p5" style="font-size:0">
                                                                                        <table border="0" width="100%" height="100%" cellpadding="0" cellspacing="0">
                                                                                            <tbody>
                                                                                                <tr>
                                                                                                    <td style="border-bottom: 1px solid #cccccc; background:none; height:1px; width:100%; margin:0px 0px 0px 0px;"></td>
                                                                                                </tr>
                                                                                            </tbody>
                                                                                        </table>
                                                                                    </td>
                                                                                </tr>
                                                                            </tbody>
                                                                        </table>
                                                                    </td>
                                                                </tr>
                                                            </tbody>
                                                        </table>
                                                    </td>
                                                </tr>
                                            </tbody>
                                        </table>
                                    </td>
                                </tr>
                            </tbody>
                        </table>
                        <table cellpadding="0" cellspacing="0" class="es-content" align="center">
                            <tbody>
                                <tr>
                                    <td class="esd-stripe" align="center">
                                        <table bgcolor="#ffffff" class="es-content-body" align="center" cellpadding="0" cellspacing="0" width="650">
                                            <tbody>
                                                <tr>
                                                    <td class="esd-structure es-p20r es-p20l" align="left">
                                                        <table cellpadding="0" cellspacing="0" width="100%">
                                                            <tbody>
                                                                <tr>
                                                                    <td width="610" class="esd-container-frame" align="center" valign="top">
                                                                        <table cellpadding="0" cellspacing="0" width="100%">
                                                                            <tbody>
                                                                                <tr>
                                                                                    <td align="center" class="esd-block-spacer es-p20" style="font-size:0">
                                                                                        <table border="0" width="100%" height="100%" cellpadding="0" cellspacing="0">
                                                                                            <tbody>
                                                                                                <tr>
                                                                                                    <td style="border-bottom: 35px solid #666666; background: none; height: 1px; width: 100%; margin: 0px;"></td>
                                                                                                </tr>
                                                                                            </tbody>
                                                                                        </table>
                                                                                    </td>
                                                                                </tr>
                                                                            </tbody>
                                                                        </table>
                                                                    </td>
                                                                </tr>
                                                            </tbody>
                                                        </table>
                                                    </td>
                                                </tr>
                                            </tbody>
                                        </table>
                                    </td>
                                </tr>
                            </tbody>
                        </table>
                        <table cellpadding="0" cellspacing="0" class="es-content" align="center">
                            <tbody>
                                <tr>
                                    <td class="esd-stripe" align="center">
                                        <table bgcolor="#ffffff" class="es-content-body" align="center" cellpadding="0" cellspacing="0" width="650">
                                            <tbody>
                                                <tr>
                                                    <td class="es-p20t es-p20r es-p20l esd-structure" align="left">
                                                        <table cellpadding="0" cellspacing="0" width="100%">
                                                            <tbody>
                                                                <tr>
                                                                    <td width="610" class="esd-container-frame" align="center" valign="top">
                                                                        <table cellpadding="0" cellspacing="0" width="100%">
                                                                            <tbody>
                                                                                <tr>
                                                                                    <td align="center" class="esd-block-text">
                                                                                        <h1><strong>Top 5 models</strong></h1>
                                                                                    </td>
                                                                                </tr>
                                                                            </tbody>
                                                                        </table>
                                                                    </td>
                                                                </tr>
                                                            </tbody>
                                                        </table>
                                                    </td>
                                                </tr>
                                            </tbody>
                                        </table>
                                    </td>
                                </tr>
                            </tbody>
                        </table>
                        <table cellpadding="0" cellspacing="0" class="es-content" align="center">
                            <tbody>
                                <tr>
                                    <td class="esd-stripe" align="center">
                                        <table bgcolor="#ffffff" class="es-content-body" align="center" cellpadding="0" cellspacing="0" width="650">
                                            <tbody>
                                                <tr>
                                                    <td class="es-p20t es-p20r es-p20l esd-structure" align="left">
                                                        <table cellpadding="0" cellspacing="0" width="100%">
                                                            <tbody>
                                                                <tr>
                                                                    <td width="610" class="esd-container-frame" align="center" valign="top">
                                                                        <table cellpadding="0" cellspacing="0" width="100%">
                                                                            <tbody>
                                                                                <tr>
                                                                                    <td align="center" class="esd-block-image" style="font-size: 0px;"><a target="_blank"><img class="adapt-img" src="{data['plot_url'][4]}" alt style="display: block;" width="600"></a></td>
                                                                                </tr>
                                                                            </tbody>
                                                                        </table>
                                                                    </td>
                                                                </tr>
                                                            </tbody>
                                                        </table>
                                                    </td>
                                                </tr>
                                            </tbody>
                                        </table>
                                    </td>
                                </tr>
                            </tbody>
                        </table>
                        <table cellpadding="0" cellspacing="0" class="es-content" align="center">
                            <tbody>
                                <tr>
                                    <td class="esd-stripe" align="center">
                                        <table bgcolor="#ffffff" class="es-content-body" align="center" cellpadding="0" cellspacing="0" width="650">
                                            <tbody>
                                                <tr>
                                                    <td class="es-p20t es-p20r es-p20l esd-structure" align="left">
                                                        <table cellpadding="0" cellspacing="0" width="100%">
                                                            <tbody>
                                                                <tr>
                                                                    <td width="610" class="esd-container-frame" align="center" valign="top">
                                                                        <table cellpadding="0" cellspacing="0" width="100%">
                                                                            <tbody>
                                                                                <tr>
                                                                                    <td align="center" class="esd-block-text">
                                                                                        <h1><strong>Categories</strong></h1>
                                                                                    </td>
                                                                                </tr>
                                                                            </tbody>
                                                                        </table>
                                                                    </td>
                                                                </tr>
                                                            </tbody>
                                                        </table>
                                                    </td>
                                                </tr>
                                            </tbody>
                                        </table>
                                    </td>
                                </tr>
                            </tbody>
                        </table>
                        <table cellpadding="0" cellspacing="0" class="es-content" align="center">
                            <tbody>
                                <tr>
                                    <td class="esd-stripe" align="center">
                                        <table bgcolor="#ffffff" class="es-content-body" align="center" cellpadding="0" cellspacing="0" width="650">
                                            <tbody>
                                                <tr>
                                                    <td class="es-p20t es-p20r es-p20l esd-structure" align="left">
                                                        <table cellpadding="0" cellspacing="0" width="100%">
                                                            <tbody>
                                                                <tr>
                                                                    <td width="610" class="esd-container-frame" align="center" valign="top">
                                                                        <table cellpadding="0" cellspacing="0" width="100%">
                                                                            <tbody>
                                                                                <tr>
                                                                                    <td align="center" class="esd-block-image" style="font-size: 0px;"><a target="_blank"><img class="adapt-img" src="{data['plot_url'][0]}" alt style="display: block;" width="600"></a></td>
                                                                                </tr>
                                                                            </tbody>
                                                                        </table>
                                                                    </td>
                                                                </tr>
                                                            </tbody>
                                                        </table>
                                                    </td>
                                                </tr>
                                            </tbody>
                                        </table>
                                    </td>
                                </tr>
                            </tbody>
                        </table>

                           <table cellpadding="0" cellspacing="0" class="es-content" align="center">
                            <tbody>
                                <tr>
                                    <td class="esd-stripe" align="center">
                                        <table bgcolor="#ffffff" class="es-content-body" align="center" cellpadding="0" cellspacing="0" width="650">
                                            <tbody>
                                                <tr>
                                                    <td class="es-p20t es-p20r es-p20l esd-structure" align="left">
                                                        <table cellpadding="0" cellspacing="0" width="100%">
                                                            <tbody>
                                                                <tr>
                                                                    <td width="610" class="esd-container-frame" align="center" valign="top">
                                                                        <table cellpadding="0" cellspacing="0" width="100%">
                                                                            <tbody>
                                                                                <tr>
                                                                                    <td align="center" class="esd-block-text">
                                                                                        <h1><strong>Families</strong></h1>
                                                                                    </td>
                                                                                </tr>
                                                                            </tbody>
                                                                        </table>
                                                                    </td>
                                                                </tr>
                                                            </tbody>
                                                        </table>
                                                    </td>
                                                </tr>
                                            </tbody>
                                        </table>
                                    </td>
                                </tr>
                            </tbody>
                        </table>
                        <table cellpadding="0" cellspacing="0" class="es-content" align="center">
                            <tbody>
                                <tr>
                                    <td class="esd-stripe" align="center">
                                        <table bgcolor="#ffffff" class="es-content-body" align="center" cellpadding="0" cellspacing="0" width="650">
                                            <tbody>
                                                <tr>
                                                    <td class="es-p20t es-p20r es-p20l esd-structure" align="left">
                                                        <table cellpadding="0" cellspacing="0" width="100%">
                                                            <tbody>
                                                                <tr>
                                                                    <td width="610" class="esd-container-frame" align="center" valign="top">
                                                                        <table cellpadding="0" cellspacing="0" width="100%">
                                                                            <tbody>
                                                                                <tr>
                                                                                    <td align="center" class="esd-block-image" style="font-size: 0px;"><a target="_blank"><img class="adapt-img" src="{data['plot_url'][3]}" width="600"></a></td>
                                                                                </tr>
                                                                            </tbody>
                                                                        </table>
                                                                    </td>
                                                                </tr>
                                                            </tbody>
                                                        </table>
                                                    </td>
                                                </tr>
                                            </tbody>
                                        </table>
                                    </td>
                                </tr>
                            </tbody>
                        </table>

                        <table cellpadding="0" cellspacing="0" class="es-content" align="center">
                            <tbody>
                                <tr>
                                    <td class="esd-stripe" align="center">
                                        <table bgcolor="#ffffff" class="es-content-body" align="center" cellpadding="0" cellspacing="0" width="650">
                                            <tbody>
                                                <tr>
                                                    <td class="es-p20t es-p20r es-p20l esd-structure" align="left">
                                                        <table cellpadding="0" cellspacing="0" width="100%">
                                                            <tbody>
                                                                <tr>
                                                                    <td width="610" class="esd-container-frame" align="center" valign="top">
                                                                        <table cellpadding="0" cellspacing="0" width="100%">
                                                                            <tbody>
                                                                                <tr>
                                                                                    <td align="center" class="esd-block-text">
                                                                                        <h1><strong>Type of request</strong></h1>
                                                                                    </td>
                                                                                </tr>
                                                                            </tbody>
                                                                        </table>
                                                                    </td>
                                                                </tr>
                                                            </tbody>
                                                        </table>
                                                    </td>
                                                </tr>
                                            </tbody>
                                        </table>
                                    </td>
                                </tr>
                            </tbody>
                        </table>
                        <table cellpadding="0" cellspacing="0" class="es-content" align="center">
                            <tbody>
                                <tr>
                                    <td class="esd-stripe" align="center">
                                        <table bgcolor="#ffffff" class="es-content-body" align="center" cellpadding="0" cellspacing="0" width="650">
                                            <tbody>
                                                <tr>
                                                    <td class="es-p20t es-p20r es-p20l esd-structure" align="left">
                                                        <table cellpadding="0" cellspacing="0" width="100%">
                                                            <tbody>
                                                                <tr>
                                                                    <td width="610" class="esd-container-frame" align="center" valign="top">
                                                                        <table cellpadding="0" cellspacing="0" width="100%">
                                                                            <tbody>
                                                                                <tr>
                                                                                    <td align="center" class="esd-block-image" style="font-size: 0px;"><a target="_blank"><img class="adapt-img" src="{data['plot_url'][2]}" alt style="display: block;" width="600"></a></td>
                                                                                </tr>
                                                                            </tbody>
                                                                        </table>
                                                                    </td>
                                                                </tr>
                                                            </tbody>
                                                        </table>
                                                    </td>
                                                </tr>
                                            </tbody>
                                        </table>
                                    </td>
                                </tr>
                            </tbody>
                        </table>

                                                <table cellpadding="0" cellspacing="0" class="es-content" align="center">
                            <tbody>
                                <tr>
                                    <td class="esd-stripe" align="center">
                                        <table bgcolor="#ffffff" class="es-content-body" align="center" cellpadding="0" cellspacing="0" width="650">
                                            <tbody>
                                                <tr>
                                                    <td class="es-p20t es-p20r es-p20l esd-structure" align="left">
                                                        <table cellpadding="0" cellspacing="0" width="100%">
                                                            <tbody>
                                                                <tr>
                                                                    <td width="610" class="esd-container-frame" align="center" valign="top">
                                                                        <table cellpadding="0" cellspacing="0" width="100%">
                                                                            <tbody>
                                                                                <tr>
                                                                                    <td align="center" class="esd-block-text">
                                                                                        <h1><strong>Clients</strong></h1>
                                                                                    </td>
                                                                                </tr>
                                                                            </tbody>
                                                                        </table>
                                                                    </td>
                                                                </tr>
                                                            </tbody>
                                                        </table>
                                                    </td>
                                                </tr>
                                            </tbody>
                                        </table>
                                    </td>
                                </tr>
                            </tbody>
                        </table>
                        <table cellpadding="0" cellspacing="0" class="es-content" align="center">
                            <tbody>
                                <tr>
                                    <td class="esd-stripe" align="center">
                                        <table bgcolor="#ffffff" class="es-content-body" align="center" cellpadding="0" cellspacing="0" width="650">
                                            <tbody>
                                                <tr>
                                                    <td class="es-p20t es-p20r es-p20l esd-structure" align="left">
                                                        <table cellpadding="0" cellspacing="0" width="100%">
                                                            <tbody>
                                                                <tr>
                                                                    <td width="610" class="esd-container-frame" align="center" valign="top">
                                                                        <table cellpadding="0" cellspacing="0" width="100%">
                                                                            <tbody>
                                                                                <tr>
                                                                                    <td align="center" class="esd-block-image" style="font-size: 0px;"><a target="_blank"><img class="adapt-img" src="{data['plot_url'][5]}" alt style="display: block;" width="600"></a></td>
                                                                                </tr>
                                                                            </tbody>
                                                                        </table>
                                                                    </td>
                                                                </tr>
                                                            </tbody>
                                                        </table>
                                                    </td>
                                                </tr>
                                            </tbody>
                                        </table>
                                    </td>
                                </tr>
                            </tbody>
                        </table>

                        <table cellpadding="0" cellspacing="0" class="es-content" align="center">
                            <tbody>
                                <tr>
                                    <td class="esd-stripe" align="center">
                                        <table bgcolor="#ffffff" class="es-content-body" align="center" cellpadding="0" cellspacing="0" width="650">
                                            <tbody>
                                                <tr>
                                                    <td class="es-p20t es-p20r es-p20l esd-structure" align="left">
                                                        <table cellpadding="0" cellspacing="0" width="100%">
                                                            <tbody>
                                                                <tr>
                                                                    <td width="610" class="esd-container-frame" align="center" valign="top">
                                                                        <table cellpadding="0" cellspacing="0" width="100%">
                                                                            <tbody>
                                                                                <tr>
                                                                                    <td align="center" class="esd-block-text">
                                                                                        <h1><strong>Markets</strong></h1>
                                                                                    </td>
                                                                                </tr>
                                                                            </tbody>
                                                                        </table>
                                                                    </td>
                                                                </tr>
                                                            </tbody>
                                                        </table>
                                                    </td>
                                                </tr>
                                            </tbody>
                                        </table>
                                    </td>
                                </tr>
                            </tbody>
                        </table>
                        <table cellpadding="0" cellspacing="0" class="es-content" align="center">
                            <tbody>
                                <tr>
                                    <td class="esd-stripe" align="center">
                                        <table bgcolor="#ffffff" class="es-content-body" align="center" cellpadding="0" cellspacing="0" width="650">
                                            <tbody>
                                                <tr>
                                                    <td class="es-p20t es-p20r es-p20l esd-structure" align="left">
                                                        <table cellpadding="0" cellspacing="0" width="100%">
                                                            <tbody>
                                                                <tr>
                                                                    <td width="610" class="esd-container-frame" align="center" valign="top">
                                                                        <table cellpadding="0" cellspacing="0" width="100%">
                                                                            <tbody>
                                                                                <tr>
                                                                                    <td align="center" class="esd-block-image" style="font-size: 0px;"><a target="_blank"><img class="adapt-img" src="{data['plot_url'][1]}" alt style="display: block;" width="600"></a></td>
                                                                                </tr>
                                                                            </tbody>
                                                                        </table>
                                                                    </td>
                                                                </tr>
                                                            </tbody>
                                                        </table>
                                                    </td>
                                                </tr>
                                            </tbody>
                                        </table>
                                    </td>
                                </tr>
                            </tbody>
                        </table>

                        <table class="esd-footer-popover es-footer" cellspacing="0" cellpadding="0" align="center">
                            <tbody>
                                <tr>
                                    <td class="esd-stripe" align="center">
                                        <table class="es-footer-body" width="650" cellspacing="0" cellpadding="0" bgcolor="#ffffff" align="center">
                                            <tbody>
                                                <tr>
                                                    <td class="esd-structure es-p20t es-p20b es-p20r es-p20l" align="left">
                                                        <table cellspacing="0" cellpadding="0" width="100%">
                                                            <tbody>
                                                                <tr>
                                                                    <td class="esd-container-frame" width="610" align="left">
                                                                        <table width="100%" cellspacing="0" cellpadding="0">
                                                                            <tbody>
                                                                                <tr>
                                                                                    <td align="center" class="esd-block-text">
                                                                                        <p style="font-family: roboto, 'helvetica neue', helvetica, arial, sans-serif;">Design e Comunica????o | Aquinos 2021</p>
                                                                                    </td>
                                                                                </tr>
                                                                            </tbody>
                                                                        </table>
                                                                    </td>
                                                                </tr>
                                                            </tbody>
                                                        </table>
                                                    </td>
                                                </tr>
                                            </tbody>
                                        </table>
                                    </td>
                                </tr>
                            </tbody>
                        </table>
                    </td>
                </tr>
            </tbody>
        </table>
    </div>
</body>

</html>


    '''

    with open(r'./_email/temp_files/email.html', 'w') as file:
            file.write(html)
            print('-------------------')
            print('email.html updated')
    
