from openpyxl import load_workbook
import xlrd
import datetime


# open the file

class DataCollector:
    
    def __init__(self, isGlobal=False):
        #Get all data from DRS DB

        file = load_workbook(filename=r'../src/assets/DRS_DB.xlsx')
        self.sheet = file['Folha1']

        max = self.sheet.max_row

        self.current_mounth_drs = {}

        x = 3
        while True:
            value = self.sheet[f'Y{x}'].value
            drs_number = self.sheet[f'A{x}'].value
            if value == None:
                break
            str_value = str(value)
            recived_month = str_value.split('-')[1].replace('0', '')
            date = datetime.datetime.now()
            past_month = date.month - 1
            drs_name = f"{self.sheet[f'A{x}'].value}_{str(self.sheet[f'B{x}'].value)}"

            #controller if is global data or pass month
            if isGlobal:
                new_dic = {drs_name:{
                        'codigo_modelo': self.sheet[f'C{x}'].value,
                        'tipologia': self.sheet[f'E{x}'].value,
                        'tipo_pedido':self.sheet[f'F{x}'].value,
                        'nome_modelo':self.sheet[f'D{x}'].value,
                        'empresa':self.sheet[f'G{x}'].value,
                        'mia':self.sheet[f'H{x}'].value,
                        'categoria':self.sheet[f'I{x}'].value,
                        'mercado':self.sheet[f'J{x}'].value,
                        'revestimentos':self.sheet[f'O{x}'].value,
                        }}
                self.current_mounth_drs.update(new_dic)

            if recived_month == str(past_month):
                new_dic = {drs_name:{
                        'codigo_modelo': self.sheet[f'C{x}'].value,
                        'tipologia': self.sheet[f'E{x}'].value,
                        'tipo_pedido':self.sheet[f'F{x}'].value,
                        'nome_modelo':self.sheet[f'D{x}'].value,
                        'empresa':self.sheet[f'G{x}'].value,
                        'mia':self.sheet[f'H{x}'].value,
                        'categoria':self.sheet[f'I{x}'].value,
                        'mercado':self.sheet[f'J{x}'].value,
                        'revestimentos':self.sheet[f'O{x}'].value,
                        }}
                self.current_mounth_drs.update(new_dic)

            x += 1
    

    def getAllDrsValues(self):
        '''Return DIC of all DRS (past month)'''
        return self.current_mounth_drs


    def getDrsNumber(self):
        '''Return nunber of all DRS (past month)'''
        return len(self.current_mounth_drs)


    def getIsMia(self):
        '''Return nunber of all MIAS (past month)'''
        data = self.current_mounth_drs

        mias = []

        for key, values in data.items():
            if values['mia']:
                mias.append(f'{key} : {values}')
        
        return len(mias)

    def getModelTipo(self, IsSorted=False):
        ''' Get all tipos of models 
            return: DIC {tipo:qnt} '''

        data = self.current_mounth_drs

        tipos = []
        tipos_dic = {}
        for key, values in data.items():
            tipo = values['tipologia']
            if tipo:
                tipos.append(tipo)
            
        for tip in tipos:
            y = {
                tip:0
            }
            tipos_dic.update(y)
        
        for x in tipos:
            tipos_dic[x] += 1

        if IsSorted:
            dic_sorted = sorted(tipos_dic.items(), key=lambda x: x[1], reverse=True)
            return dic_sorted 
        else:
            return  tipos_dic

        return tipos_dic
    

    def getMarkets(self, IsSorted=False):
        ''' Get all markets 
            return: DIC {tipo:qnt} '''

        data = self.current_mounth_drs

        tipos = []
        tipos_dic = {}
        for key, values in data.items():
            tipo = values['mercado']
            if tipo:
                tipos.append(tipo)
            
        for tip in tipos:
            y = {
                tip:0
            }
            tipos_dic.update(y)
        
        for x in tipos:
            tipos_dic[x] += 1

        if IsSorted:
            dic_sorted = sorted(tipos_dic.items(), key=lambda x: x[1], reverse=True)
            return dic_sorted 
        else:
            return  tipos_dic

        return tipos_dic


    def getCategory(self, IsSorted=False):
        ''' Get all markets return: DIC {tipo:qnt} '''

        data = self.current_mounth_drs

        tipos = []
        tipos_dic = {}
        for key, values in data.items():
            tipo = values['categoria']
            if tipo:
                tipos.append(tipo)
            
        for tip in tipos:
            y = {
                tip:0
            }
            tipos_dic.update(y)
        
        for x in tipos:
            tipos_dic[x] += 1

        if IsSorted:
            dic_sorted = sorted(tipos_dic.items(), key=lambda x: x[1], reverse=True)
            return dic_sorted 
        else:
            return  tipos_dic

        return tipos_dic


    def getRequestTypes(self, IsSorted=False):
        ''' Get all markets 
            return: DIC {tipo:qnt} '''

        data = self.current_mounth_drs

        tipos = []
        tipos_dic = {}
        for key, values in data.items():
            tipo = values['tipo_pedido']
            if tipo:
                xtipo = tipo.split(',')
                for j in xtipo:
                    if len(j) > 1:
                        tipos.append(j.strip())
                    
        for tip in tipos:
            y = {
                tip:0
            }
            tipos_dic.update(y)
        
        for x in tipos:
            tipos_dic[x] += 1

        if IsSorted:
            dic_sorted = sorted(tipos_dic.items(), key=lambda x: x[1], reverse=True)
            return dic_sorted 
        else:
            return  tipos_dic

    
    def getFabrics(self, IsSorted=False):
        ''' Get all markets 
            return: DIC {tipo:qnt} '''

        data = self.current_mounth_drs

        tipos = []
        tipos_dic = {}
        for key, values in data.items():
            tipo = values['revestimentos']
            if tipo:
                xtipo = tipo.split(',')
                for j in xtipo:
                    if len(j) > 1:
                        tipos.append(j.strip())
                    
        for tip in tipos:
            y = {
                tip:0
            }
            tipos_dic.update(y)
        
        for x in tipos:
            tipos_dic[x] += 1
    

        if IsSorted:
            dic_sorted = sorted(tipos_dic.items(), key=lambda x: x[1], reverse=True)
            return dic_sorted 
        else:
            return  tipos_dic
            
                


